import json
import os.path

import click
from openpyxl import load_workbook

#todo remove many blanks

headers = ['PRONOUNS', 'DEMOGRAPHICS', 'TEMPORAL-DISTANCE STUDY',
           'SPATIAL-DISTANCE STUDY', 'SOCIAL-DISTANCE STUDY', 'LIKELIHOOD STUDY',
           'FOLLOW-UP QUESTIONS (AHS-scale, PANAS-scale, manipulation checks)',
           'MANIPULATION CHECKS', 'COUNTRIES', 'Analysis-Holism Scale (AHS)',
           'Positive and Negative Affect Schedule (PANAS)']

def parse_xlsx(xlsx, settings):
    """
    Parses xslx file
    :param xlsx: xslx file
    :return: parsed values as dict
    """
    with open(settings, 'r') as sf:
        settings = json.loads(sf.read())
        wb = load_workbook(xlsx)
        texts = wb['Texts']
        #data = {"em": {}}
        data = settings

        for _ids, _en, _text, _, _, _styling in texts.iter_rows(min_row=2, max_col=6, values_only=False):
            ids = _ids.value
            if ids:
                en = _en.value
                text = _text.value
                styling = _styling.value
                bold = _text.font.b
                italic = _text.font.i
                ids_list = ids.split(';')
                if text:
                    if styling:

                        if styling == 'b':
                            text = f"<strong>{text}</strong>"

                        elif styling == 'bc':
                            text = f'<div style="text-align: center;"><strong>{text}</strong></div>'

                        elif styling == 'c':
                            text = f'<div style="text-align: center;">{text}</div>'
                    if not '<table' in text:

                        text = text.replace('\n', '<br/>')
                    else:
                        print(text)
                else:
                    text = ''

                for i in ids_list:
                    if i in headers:
                        continue

                    id, qtype, study = i.split('_')



                    if qtype != "QuestionText" and qtype != "EM":
                        if qtype.startswith('Choice'):
                            optionno = qtype.split('Choice')[1]
                            qtype = "Choice"
                        else:
                            optionno = qtype.split('Answer')[1]
                            qtype = "Answer"


                    if id not in data[study]['data']:
                        if qtype == "QuestionText":
                            data[study]['data'][id] = {'text': text, "options": {}, "answers":{}}
                        elif qtype == "Choice":
                            data[study]['data'][id] = {'options': {optionno: {'text': text}}}
                        elif qtype == "Answer":
                            data[study]['data'][id] = {'answers': {optionno: {'text': text}}}
                        elif qtype == "EM":
                            data[study]['data']['em'][id.replace('#', '_')] = text


                    else:
                        if qtype == "Choice":
                            data[study]['data'][id]['options'][optionno] = {'text': text}

                        elif qtype == "Answer":
                            data[study]['data'][id]['answers'][optionno] = {'text': text}

                        elif qtype == "QuestionText":
                            data[study]['data'][id]['text'] = text

    return data


def get_embedded(q, data, translations):
    if 'Flow' in data:
        for x in data['Flow']:
            if x['Type'] == 'EmbeddedData':
                flowId = x['FlowID']
                if 'Value' in x['EmbeddedData'][0]:
                    if flowId in translations['em']:
                        x['EmbeddedData'][0]['Value'] = translations['em'][flowId]
            get_embedded(q, x, translations)

    return q



@click.command()
@click.option('-x', 'excel_path', required=True, help="Path to translation excel file")
@click.option('-l', 'language', required=True, help="Survey target language code. e.g. SV, EN")
@click.option('-s', 'settings', required=True, help="path to settings.json")
def main(language, excel_path ,settings):
    translations = parse_xlsx(excel_path, settings)
    with open('all_data.json','w') as o:
        o.write(json.dumps(translations, indent=4))


    for key,value in translations.items():
        print("SURVEY",key)
        basename = os.path.basename(value['path'])
        new_name = basename.replace('_en_template',f'_{language.lower()}')
        if not os.path.exists(language.lower()):
            os.makedirs(language.lower())


        with open(value['path'], encoding='utf-8') as f:
            qsf = json.loads(f.read())
            qsf['SurveyEntry']['SurveyLanguage'] = language.upper()

            for q in qsf['SurveyElements']:
                if q['Element'] == "SQ":
                    id = q['PrimaryAttribute']
                    if id in translations[key]['data']:

                        q['Payload']['QuestionText'] = translations[key]['data'][id]['text']

                        if "Choices" in q['Payload']:
                            if 'exclude' in translations[key] and id in translations[key]['exclude']:
                                continue
                            for k, v in q['Payload']["Choices"].items():
                                try:
                                    q['Payload']["Choices"][k]['Display'] = translations[key]['data'][id]['options'][k]['text']
                                except:
                                    print(id)
                                    print(q['Payload']["Choices"][k]['Display'])
                                    raise
                        if "Answers" in q['Payload']:
                            if 'exclude' in translations[key] and id in translations[key]['exclude']:
                                continue

                            for k, v in q['Payload']["Answers"].items():
                                try:
                                    #print(translations[key]['data'][id])
                                    q['Payload']["Answers"][k]['Display'] = translations[key]['data'][id]['answers'][k]['text']
                                except:
                                    print(id)
                                    #print(q['Payload']["Answers"][k]['Display'])
                                    raise

                            #print(q['Payload']["Answers"])


                elif q['Element'] == "FL":
                    embedded = get_embedded(q, q['Payload'], translations[key]['data'])
                    q = embedded

            with open(os.path.join(language.lower(),new_name), 'w', encoding='utf-8') as of:
                of.write(json.dumps(qsf, indent=4, ensure_ascii=False))


if __name__ == '__main__':
    main()
