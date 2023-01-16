import csv
import hashlib
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Alignment

wb = load_workbook('data/CLIMR_master_Likelihood_BIF.xlsx')
sheet = wb['Texts']
fontStyle = Font(name="Calibri", size=12, color=colors.BLACK)
dest_filename = 'data/CSV_test_BIF.xlsx'
data = {}

with open('data/ANDREAS_CLIMR_English_Social_BIF - Copy-EN-SV(3).csv', 'r', encoding='utf8') as f:
    reader = csv.reader(f)
    header = reader.__next__()
    survey_id = reader.__next__()

    for id, en, ver in reader:
        mod_str = en.lower().replace(' ', '')
        md5 = hashlib.md5(mod_str.encode('utf-8')).hexdigest()
        if md5 not in data:
            data[md5] = {"text":en,"ids":[id]}
        else:
            data[md5]['ids'].append(id)

row = 1

for k,v in data.items():
    row += 1
    for col in range(1, 2):
        if v['ids'][0] == 'SurveyTitle' or v['ids'][0] == 'SurveyDescription':
            continue
        a = sheet.cell(column=col, row=row, value=";".join(v['ids']))
        b = sheet.cell(column=col+1, row=row, value=f"{v['text']}")
        a.font = fontStyle
        b.font = fontStyle
        b.alignment = Alignment(wrapText=True)

wb.save(filename = dest_filename)