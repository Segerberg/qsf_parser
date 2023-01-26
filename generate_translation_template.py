import hashlib
import json

import click
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection, colors


data = {}
embedded = {}
blockorder = {}
ordered_data = {}


def get_embedded(data):
    """
    Function for recursively yield embedded data values
    :param data: payload from qsf FL survey element
    :return: Generator
    """

    if 'Flow' in data:
        for k in data['Flow']:
            if k['Type'] == 'EmbeddedData':
                flowId = f"{k['FlowID'].replace('_','#')}_EM"
                if 'Value' not in k['EmbeddedData'][0]:
                    continue
                elif k['EmbeddedData'][0]['Value'].startswith('$'):
                    continue
                q_text = k['EmbeddedData'][0]['Value']
                mod_str = q_text.lower().replace(' ', '').replace('\n', '')
                md5 = hashlib.md5(mod_str.encode('utf-8')).hexdigest()
                if md5 not in embedded:
                    embedded[md5] = {'text':q_text, 'ids':[flowId]}
                else:
                    embedded[md5]['ids'].append(flowId)

            get_embedded(k)


@click.command()
@click.option('-q','qsf_path', required=True)
@click.option('-o', 'output', required=True)
@click.option('-s', '--striphtml', is_flag=True)
def main(qsf_path, output, striphtml):
    wb = Workbook()
    info = wb['Sheet']
    info.title = 'Info'
    sheet = wb.create_sheet(title="Texts")
    sheet['A1'] = 'ID'
    sheet['B1'] = 'English'
    sheet['C1'] = 'Translation'
    fontStyle = Font(name="Calibri", size=12, color=colors.BLACK)

    with open(qsf_path, 'r', encoding='utf8') as f:
        qsf = json.loads(f.read())

        for q in qsf['SurveyElements']:
            if q['Element'] == "FL":
                get_embedded(q['Payload'])

        for q in qsf['SurveyElements']:
            if q['Element'] == "BL":
                for k, v in q['Payload'].items():
                    if v['Type'] == "Default" or v['Type'] == 'Standard':
                        for blq in v['BlockElements']:
                            if blq['Type'] == "Question":
                                blockorder[blq['QuestionID']] = k

        for q in qsf['SurveyElements']:
            if q['Element'] == "SQ":
                if q['PrimaryAttribute'] in blockorder:
                    order = blockorder[q['PrimaryAttribute']]

                id = f"{q['PrimaryAttribute']}_QuestionText"
                q_text = q['Payload']['QuestionText']
                mod_str = q_text.lower().replace(' ', '').replace('\n', '')
                md5 = hashlib.md5(mod_str.encode('utf-8')).hexdigest()

                if md5 not in data:
                    data[md5] = {"order": order, "text": q_text, "ids": [id]}

                else:
                    data[md5]['ids'].append(id)

                if "Choices" in q['Payload']:
                    for k, v in q['Payload']['Choices'].items():
                        c_text = v['Display']
                        id = f"{q['PrimaryAttribute']}_Choice{k}"
                        mod_str = c_text.lower().replace(' ', '')
                        md5 = hashlib.md5(mod_str.encode('utf-8')).hexdigest()
                        if md5 not in data:
                            data[md5] = {"order": order, "text": c_text, "ids": [id]}
                        else:
                            data[md5]['ids'].append(id)

        for k, v in data.items():
            if int(v['order']) not in ordered_data:
                ordered_data[int(v['order'])] = {k: v}
            else:
                ordered_data[int(v['order'])].update({k: v})

    keys = sorted(ordered_data)
    sorted_data = {i: ordered_data[i] for i in keys}
    sorted_data.update({"em": embedded})

    row = 1
    block_count = 1

    for key, value in sorted_data.items():

        row += 1
        for col in range(1, 4):
            block = sheet.cell(column=col, row=row, value="#")
            block.fill = PatternFill("solid", fgColor="fcba03")
            block.font = fontStyle
        block_count += 1

        for k, v in value.items():
            row += 1

            a = sheet.cell(column=1, row=row, value=";".join(v['ids']))
            a.protection = Protection(locked=True, hidden=True)
            a.font = fontStyle
            a.alignment = Alignment(vertical='top')

            cell_text = v['text']
            if striphtml:
                soup = BeautifulSoup(v['text'], 'html.parser')
                cell_text = str(soup.get_text())

            b = sheet.cell(column=2, row=row, value=f"{cell_text}")
            b.protection = Protection(locked=False)
            b.font = fontStyle
            b.alignment = Alignment(wrapText=True, vertical='top')

            c = sheet.cell(column=3, row=row)
            c.protection = Protection(locked=False)
            c.font = fontStyle
            c.alignment = Alignment(wrapText=True, vertical='top')

    wb.save(filename=output)


if __name__ == '__main__':
    main()
