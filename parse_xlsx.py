import json

import click
from openpyxl import load_workbook


def parse_xlsx(xlsx):
    """
    Parses xslx file
    :param xlsx: xslx file
    :return: parsed values as dict
    """
    wb = load_workbook(xlsx)
    texts = wb['Texts']
    data = {"em": {}}
    for _ids, _en, _text in texts.iter_rows(min_row=2, max_col=3, values_only=False):
        ids = _ids.value
        en = _en.value
        text = _text.value

        bold = _text.font.b
        italic = _text.font.i
        ids_list = ids.split(';')
        for i in ids_list:
            if '#' in i:  # todo
                continue
            id, qtype = i.split('_')

            if text:

                if bold and italic:
                    text = f"<strong><em>{text}<em></strong>"
                elif bold:
                    text = f"<strong>{text}</strong>"
                elif italic:
                    text = f"<em>{text}</em>"

                text = text.replace('\n', '<br/>')
            else:
                text = ''

            if qtype != "QuestionText" and qtype != "EM":
                optionno = qtype.split('Choice')[1]
                qtype = "Choice"

            if id not in data:
                if qtype == "QuestionText":

                    data[id] = {'text': text, "options": {}}
                elif qtype == "Choice":
                    data[id] = {'options': {optionno: {'text': text}}}
                elif qtype == "EM":
                    data['em'][id.replace('#', '_')] = text

            else:
                if qtype == "Choice":
                    data[id]['options'][optionno] = {'text': text}
                elif qtype == "QuestionText":
                    data[id]['text'] = text

    return data


def get_embedded(data, translations):
    """
    Function for recursively yield embedded data values
    :param data: payload from qsf FL survey element
    :param translations: Look up dict
    :return: Generator
    """
    if 'Flow' in data:
        for k in data['Flow']:
            if k['Type'] == 'EmbeddedData':
                flowId = k['FlowID']
                if 'Value' in k['EmbeddedData'][0]:
                    if flowId in translations['em']:
                        k['EmbeddedData'][0]['Value'] = translations['em'][flowId]
                        yield k
            yield from get_embedded(k, translations)


@click.command()
@click.option('-q', 'qsf_path', required=True, help="Path to input qsf")
@click.option('-x', 'excel_path', required=True, help="Path to translation excel file")
@click.option('-o', 'output', default="output.qsf", help="qsf filename to be generated")
def main(qsf_path, excel_path, output):
    translations = parse_xlsx(excel_path)
    with open(qsf_path) as f:
        qsf = json.loads(f.read())
        for q in qsf['SurveyElements']:
            if q['Element'] == "SQ":
                id = q['PrimaryAttribute']
                if id in translations:
                    q['Payload']['QuestionText'] = translations[id]['text']
                    if "Choices" in q['Payload']:
                        for k, v in q['Payload']["Choices"].items():
                            q['Payload']["Choices"][k]['Display'] = translations[id]['options'][k]

            elif q['Element'] == "FL":
                embedded = get_embedded(q['Payload'], translations)

                for e in embedded:
                    q['Payload'].update(e)

        with open(output, 'w', encoding='utf-8') as of:
            of.write(json.dumps(qsf, indent=4, ensure_ascii=False))


if __name__ == '__main__':
    main()
